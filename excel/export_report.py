"""
export_report.py
Generates a formatted Excel workbook (openpyxl) with:
 - Raw data sheet
 - Summary pivot-style tables
 - A native Excel chart

Run:
    python excel/export_report.py

Output:
    excel/student_performance_report.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

df = pd.read_csv("data/students.csv")

wb = Workbook()

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")


def autofit(ws, ncols):
    for col in range(1, ncols + 1):
        letter = ws.cell(row=1, column=col).column_letter
        max_len = max(
            [len(str(ws.cell(row=r, column=col).value)) for r in range(1, ws.max_row + 1)]
        )
        ws.column_dimensions[letter].width = min(max_len + 3, 35)


# --- Sheet 1: Raw data ---
ws1 = wb.active
ws1.title = "Raw Data"
for row in dataframe_to_rows(df, index=False, header=True):
    ws1.append(row)
style_header(ws1, df.shape[1])
autofit(ws1, df.shape[1])
ws1.freeze_panes = "A2"

# --- Sheet 2: Summary by parent education ---
ws2 = wb.create_sheet("Summary - Parent Ed")
summary1 = (
    df.groupby("parent_education")
    .agg(n_students=("student_id", "count"), avg_grade=("final_grade_pct", "mean"), pass_rate=("passed", "mean"))
    .reset_index()
)
summary1["avg_grade"] = summary1["avg_grade"].round(2)
summary1["pass_rate"] = (summary1["pass_rate"] * 100).round(1)
for row in dataframe_to_rows(summary1, index=False, header=True):
    ws2.append(row)
style_header(ws2, summary1.shape[1])
autofit(ws2, summary1.shape[1])

chart = BarChart()
chart.title = "Average Final Grade by Parent Education"
chart.y_axis.title = "Average Grade (%)"
chart.x_axis.title = "Parent Education"
data = Reference(ws2, min_col=3, min_row=1, max_row=summary1.shape[0] + 1)
cats = Reference(ws2, min_col=1, min_row=2, max_row=summary1.shape[0] + 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws2.add_chart(chart, "F2")

# --- Sheet 3: Summary by test prep / tutoring ---
ws3 = wb.create_sheet("Summary - Prep & Tutoring")
summary2 = (
    df.groupby(["test_preparation", "tutoring"])
    .agg(n_students=("student_id", "count"), avg_grade=("final_grade_pct", "mean"))
    .reset_index()
)
summary2["avg_grade"] = summary2["avg_grade"].round(2)
for row in dataframe_to_rows(summary2, index=False, header=True):
    ws3.append(row)
style_header(ws3, summary2.shape[1])
autofit(ws3, summary2.shape[1])

# --- Sheet 4: KPIs ---
ws4 = wb.create_sheet("KPIs")
kpis = [
    ("Total Students", len(df)),
    ("Overall Pass Rate (%)", round(df.passed.mean() * 100, 2)),
    ("Average Final Grade (%)", round(df.final_grade_pct.mean(), 2)),
    ("Average Study Hours/Week", round(df.study_hours_week.mean(), 2)),
    ("Average Attendance Rate (%)", round(df.attendance_rate.mean() * 100, 2)),
]
ws4.append(["Metric", "Value"])
for k, v in kpis:
    ws4.append([k, v])
style_header(ws4, 2)
autofit(ws4, 2)

wb.save("excel/student_performance_report.xlsx")
print("Saved excel/student_performance_report.xlsx")
